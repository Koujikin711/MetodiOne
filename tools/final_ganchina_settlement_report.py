import calendar
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

SRC = r"C:\Users\nikit\Downloads\ОСВ Клиника (1).xlsx"
OUT = r"C:\Users\nikit\Desktop\Отчет_выручка_дебет_кредит_Ганчина.xlsx"

REPORT_START = date(2026, 1, 1)
REPORT_END = date(2026, 4, 30)
GANCHINA_BASE_MONTHLY = 30000.0
GANCHINA_SHARE = 0.40


@dataclass
class Contract:
    contract_id: int
    client: str
    item: str
    contract_price: float
    period_raw: str
    period_days: int
    start_date: date
    end_date: date
    paid: float = 0.0
    rows: list[dict] = field(default_factory=list)


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def num(v):
    try:
        if v is None or v == "":
            return 0.0
        return float(v)
    except Exception:
        return 0.0


def norm(v) -> str:
    return "" if v is None else str(v).strip()


def parse_period_days(period_raw, item_name=""):
    s = norm(period_raw).lower()
    m = re.search(r"(\d+)", s)
    n = int(m.group(1)) if m else 1
    if n <= 2:
        return 14  # новое правило пользователя: 1-2 дня считаем как минимум 14
    if "мес" in s or "месяц" in s:
        return max(1, n) * 30
    if n == 3 and "курс" in norm(item_name).lower():
        return 90
    return max(1, n)


def format_period_norm(days: int) -> str:
    if days % 30 == 0 and days >= 30:
        months = days // 30
        return f"{months} месяца" if months in (2, 3, 4) else f"{months} месяцев"
    return f"{days} дней"


def month_key(d: date) -> str:
    return d.strftime("%Y-%m")


def iter_month_starts(start: date, end: date):
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        yield date(y, m, 1)
        if m == 12:
            y += 1
            m = 1
        else:
            m += 1


def month_end(d: date) -> date:
    return date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])


def overlap_days(a0: date, a1: date, b0: date, b1: date) -> int:
    lo = max(a0, b0)
    hi = min(a1, b1)
    if hi < lo:
        return 0
    return (hi - lo).days + 1


def read_rows(ws):
    rows = []
    for r in range(4, ws.max_row + 1):
        d = as_date(ws.cell(r, 1).value)
        if not d or d < REPORT_START or d > REPORT_END:
            continue
        rows.append(
            {
                "date": d,
                "contract_price": num(ws.cell(r, 2).value),
                "period_raw": norm(ws.cell(r, 3).value),
                "revenue": num(ws.cell(r, 4).value),
                "expense": num(ws.cell(r, 5).value),
                "client": norm(ws.cell(r, 8).value),
                "item": norm(ws.cell(r, 11).value),
                "article": norm(ws.cell(r, 12).value),
            }
        )
    rows.sort(key=lambda x: (x["date"], x["client"], x["item"]))
    return rows


def is_service_income(x: dict) -> bool:
    return x["revenue"] > 0 and x["contract_price"] > 0 and bool(x["client"])


def build_contracts(service_rows: list[dict]) -> list[Contract]:
    contracts: list[Contract] = []
    open_contracts: dict[tuple[str, float, str], list[Contract]] = defaultdict(list)
    cid = 1

    for x in service_rows:
        key = (x["client"].lower(), round(x["contract_price"], 2), x["item"].lower())
        left = x["revenue"]

        for c in open_contracts[key]:
            if left <= 0:
                break
            rem = max(c.contract_price - c.paid, 0.0)
            if rem <= 1e-9:
                continue
            take = min(rem, left)
            c.paid += take
            c.rows.append({**x, "allocated": take})
            left -= take

        while left > 1e-9:
            p_days = parse_period_days(x["period_raw"], x["item"])
            c = Contract(
                contract_id=cid,
                client=x["client"],
                item=x["item"],
                contract_price=x["contract_price"],
                period_raw=x["period_raw"],
                period_days=p_days,
                start_date=x["date"],
                end_date=x["date"] + timedelta(days=p_days - 1),
            )
            cid += 1
            take = min(c.contract_price, left)
            c.paid += take
            c.rows.append({**x, "allocated": take})
            left -= take
            contracts.append(c)
            open_contracts[key].append(c)

    return contracts


def earned_to_date(c: Contract, at_date: date) -> float:
    used = overlap_days(c.start_date, c.end_date, c.start_date, min(c.end_date, at_date))
    ratio = min(1.0, max(0.0, used / c.period_days))
    return c.contract_price * ratio


def ganchina_base_for_month(mk: str) -> float:
    # По уточнению: фикс за март и апрель не начисляется.
    if mk in ("2026-03", "2026-04"):
        return 0.0
    return GANCHINA_BASE_MONTHLY


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["ОСВ"]

    rows = read_rows(ws)
    service_rows = [x for x in rows if is_service_income(x)]
    contracts = build_contracts(service_rows)

    months = [m for m in iter_month_starts(REPORT_START, REPORT_END)]
    stats = {
        month_key(m): {
            "earned": 0.0,
            "paid": 0.0,
            "expense": 0.0,
            "debit": 0.0,
            "credit": 0.0,
            "net_profit": 0.0,
            "g_base": ganchina_base_for_month(month_key(m)),
            "g_share": 0.0,
            "g_due": 0.0,
            "g_paid": 0.0,
        }
        for m in months
    }

    for x in rows:
        mk = month_key(x["date"])
        if mk in stats:
            stats[mk]["paid"] += x["revenue"]
            stats[mk]["expense"] += x["expense"]
            # факт выплат Ганчине
            if x["expense"] > 0 and "ганчина" in x["client"].lower() and "фот" in x["article"].lower():
                stats[mk]["g_paid"] += x["expense"]

    for c in contracts:
        daily = c.contract_price / c.period_days
        for m in months:
            m0, m1 = m, month_end(m)
            d = overlap_days(c.start_date, c.end_date, m0, m1)
            if d > 0:
                stats[month_key(m)]["earned"] += daily * d

    contract_rows = []
    total_debit = total_credit = total_earned = total_paid = 0.0

    for c in contracts:
        earned = earned_to_date(c, REPORT_END)
        paid = c.paid
        ended = c.end_date <= REPORT_END
        debit = max(earned - paid, 0.0) if ended else 0.0
        credit = max(paid - earned, 0.0) if not ended else max(paid - c.contract_price, 0.0)

        total_debit += debit
        total_credit += credit
        total_earned += earned
        total_paid += paid

        contract_rows.append(
            {
                "id": c.contract_id,
                "client": c.client,
                "item": c.item,
                "start": c.start_date,
                "end": c.end_date,
                "period_raw": c.period_raw,
                "period_norm": format_period_norm(c.period_days),
                "price": c.contract_price,
                "earned": earned,
                "paid": paid,
                "debit": debit,
                "credit": credit,
            }
        )

    for mk, st in stats.items():
        st["net_profit"] = st["earned"] - st["expense"]
        diff = st["earned"] - st["paid"]
        st["debit"] = diff if diff > 0 else 0.0
        st["credit"] = -diff if diff < 0 else 0.0
        base_for_share = st["net_profit"] - st["debit"]  # новое правило
        st["g_share"] = max(0.0, GANCHINA_SHARE * base_for_share)
        st["g_due"] = st["g_base"] + st["g_share"]

    g_due_total = sum(v["g_due"] for v in stats.values())
    g_paid_total = sum(v["g_paid"] for v in stats.values())
    g_delta = g_paid_total - g_due_total

    out = Workbook()
    sh = out.active
    sh.title = "Итог"
    sh["A1"] = "Отчет по выручке/дебету/кредиту и взаиморасчетам с Ганчиной"
    sh["A1"].font = Font(bold=True, size=13)

    sh["A3"] = "Период"
    sh["B3"] = f"{REPORT_START.isoformat()} .. {REPORT_END.isoformat()}"
    sh["A5"] = "Начисленная выручка (по сроку)"
    sh["B5"] = total_earned
    sh["A6"] = "Фактические оплаты"
    sh["B6"] = total_paid
    sh["A7"] = "Дебет (срок прошел, недоплата)"
    sh["B7"] = total_debit
    sh["A8"] = "Кредит (переплата до срока)"
    sh["B8"] = total_credit

    sh["A10"] = "Ганчина должна была получить (база+40%*(чист.прибыль-дебиторка))"
    sh["B10"] = g_due_total
    sh["A11"] = "Ганчина фактически получила"
    sh["B11"] = g_paid_total
    sh["A12"] = "Разница (факт - должна)"
    sh["B12"] = g_delta
    sh["A13"] = "Итог"
    sh["B13"] = (
        "Переплата: Ганчина должна компании" if g_delta > 0 else "Недоплата: компания должна Ганчине" if g_delta < 0 else "Взаиморасчеты закрыты"
    )

    for rr in (5, 6, 7, 8, 10, 11, 12):
        sh[f"B{rr}"].number_format = "#,##0.00"
    sh.column_dimensions["A"].width = 82
    sh.column_dimensions["B"].width = 36

    ms = out.create_sheet("Помесячно")
    ms.append([
        "Месяц",
        "Начисленная выручка",
        "Фактическая оплата",
        "Расход",
        "Чистая прибыль",
        "Дебиторка (дебет)",
        "Кредит",
        "База Ганчина",
        "40% от (чист.прибыль-дебиторка)",
        "Должны Ганчине",
        "Оплачено Ганчине",
        "Разница",
    ])
    for c in ms[1]:
        c.font = Font(bold=True)

    for mk in sorted(stats):
        st = stats[mk]
        ms.append(
            [
                mk,
                st["earned"],
                st["paid"],
                st["expense"],
                st["net_profit"],
                st["debit"],
                st["credit"],
                st["g_base"],
                st["g_share"],
                st["g_due"],
                st["g_paid"],
                st["g_paid"] - st["g_due"],
            ]
        )

    for row in ms.iter_rows(min_row=2, max_row=ms.max_row, min_col=2, max_col=12):
        for cell in row:
            cell.number_format = "#,##0.00"
    for col, w in zip("ABCDEFGHIJKL", [10, 18, 18, 14, 14, 15, 12, 12, 24, 16, 16, 14]):
        ms.column_dimensions[col].width = w

    cd = out.create_sheet("Договоры")
    cd.append([
        "ID",
        "Клиент",
        "Товар/услуга",
        "Начало",
        "Окончание",
        "Период",
        "Стоимость (B)",
        "Начислено",
        "Оплачено",
        "Дебет",
        "Кредит",
    ])
    for c in cd[1]:
        c.font = Font(bold=True)

    for x in sorted(contract_rows, key=lambda r: (r["start"], r["client"], r["id"])):
        cd.append([
            x["id"],
            x["client"],
            x["item"],
            x["start"].isoformat(),
            x["end"].isoformat(),
            x["period_norm"],
            x["price"],
            x["earned"],
            x["paid"],
            x["debit"],
            x["credit"],
        ])

    for row in cd.iter_rows(min_row=2, max_row=cd.max_row, min_col=7, max_col=11):
        for cell in row:
            cell.number_format = "#,##0.00"
    for col, w in zip("ABCDEFGHIJK", [7, 28, 24, 12, 12, 14, 16, 16, 16, 14, 14]):
        cd.column_dimensions[col].width = w

    out.save(OUT)
    print(OUT)
    print(f"TOTAL_EARNED={total_earned:.2f}")
    print(f"TOTAL_PAID={total_paid:.2f}")
    print(f"TOTAL_DEBIT={total_debit:.2f}")
    print(f"TOTAL_CREDIT={total_credit:.2f}")
    print(f"GANCHINA_DUE={g_due_total:.2f}")
    print(f"GANCHINA_PAID={g_paid_total:.2f}")
    print(f"GANCHINA_DELTA={g_delta:.2f}")


if __name__ == "__main__":
    main()
