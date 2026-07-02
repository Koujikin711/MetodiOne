import asyncio
import csv
import io
import json
import logging
import secrets
from typing import Annotated, Any

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, Request, UploadFile, status
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from starlette.responses import Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.services.tariff import count_company_integrations
from app.services.tariff_effective import effective_tariff_max_integrations
from app.core.deps import CurrentCompanyId, CurrentUser
from app.database import get_db
from app.services.green_incoming import parse_green_message_data
from app.services.green_api_send import send_green_text_async
from app.services.whatsapp_automation import send_welcome_if_first_incoming
from app.services.green_api_settings import (
    fetch_green_settings,
    fetch_green_state_instance,
    green_api_base_from_config,
    push_green_incoming_webhook,
    resolve_public_api_base,
)
from app.services.google_sheets_sync import sync_google_sheet_integration
from app.services.green_api_backfill import sync_green_api_backfill
from app.services.green_incoming_media import persist_incoming_green_media_if_needed
from app.services.instagram_webhook import handle_instagram_webhook, meta_hub_challenge_response
from app.services.integration_inbound import (
    add_incoming_message as _add_incoming_message,
    find_incoming_message_by_provider_id as _find_incoming_message_by_provider_id,
    create_lead_from_integration,
    norm_phone as _norm_phone,
    upsert_thread as _upsert_thread,
)
from datetime import UTC, datetime

from app.models import (
    ChatThread,
    Integration,
    IntegrationProvider,
    Lead,
    Pipeline,
    PipelineStage,
    UserRole,
)
from app.services.chief_expert_access import assert_owner_or_chief_expert
from app.schemas.integrations import IntegrationCreate, IntegrationRead, IntegrationUpdate
from app.schemas.lead import LeadRead

router = APIRouter(prefix="/integrations", tags=["integrations"])
logger = logging.getLogger(__name__)

_SECRET_CFG_KEYS = frozenset(
    {
        "api_token",
        "apiToken",
        "apiTokenInstance",
        "page_access_token",
        "pageAccessToken",
        "app_secret",
        "appSecret",
        "app_password",
        "appPassword",
    },
)


def _normalize_webhook_token(raw: str | None) -> str | None:
    """Убирает префиксы Bearer/Basic — Green API иногда шлёт их дважды."""
    if not raw or not str(raw).strip():
        return None
    s = str(raw).strip()
    for _ in range(3):
        low = s.lower()
        if low.startswith("bearer "):
            s = s[7:].strip()
            continue
        if low.startswith("basic "):
            s = s[6:].strip()
            continue
        break
    return s or None


def _webhook_token_matches(provided: str | None, expected: str) -> bool:
    """compare_digest падает ValueError при разной длине строк — для вебхука это 400/500 вместо 403."""
    got = _normalize_webhook_token(provided)
    exp = _normalize_webhook_token(expected)
    if not got or not exp:
        return False
    if len(got) != len(exp):
        return False
    return secrets.compare_digest(got, exp)


def _token_from_authorization_header(raw: str | None) -> str | None:
    """Green API шлёт webhookUrlToken в заголовке Authorization: Bearer <token>."""
    return _normalize_webhook_token(raw)


def _integration_read(row: Integration, *, setup_note: str | None = None) -> IntegrationRead:
    cfg = row.config
    has_token = False
    safe_cfg: dict | None = None
    if cfg and isinstance(cfg, dict):
        t = (
            cfg.get("api_token")
            or cfg.get("apiToken")
            or cfg.get("apiTokenInstance")
            or cfg.get("page_access_token")
            or cfg.get("pageAccessToken")
            or cfg.get("app_password")
            or cfg.get("appPassword")
        )
        has_token = bool(t)
        safe_cfg = {k: v for k, v in cfg.items() if k not in _SECRET_CFG_KEYS}
        if not safe_cfg and not has_token:
            safe_cfg = None
        elif not safe_cfg:
            safe_cfg = {}
    pv = row.provider.value if hasattr(row.provider, "value") else str(row.provider)
    return IntegrationRead(
        id=row.id,
        name=row.name,
        provider=pv,
        is_active=row.is_active,
        pipeline_id=row.pipeline_id,
        stage_id=row.stage_id,
        manager_close_deal_enabled=bool(getattr(row, "manager_close_deal_enabled", False)),
        config=safe_cfg,
        has_api_token=has_token,
        setup_note=setup_note,
    )


def _merge_instagram_config(old: dict | None, new: dict | None) -> dict:
    merged = dict(old or {})
    if not new:
        return merged
    skip_empty = ("page_access_token", "pageAccessToken", "app_secret", "appSecret")
    for k, v in new.items():
        if k in skip_empty and (v is None or str(v).strip() == ""):
            continue
        merged[k] = v
    return merged


def _merge_gmail_config(old: dict | None, new: dict | None) -> dict:
    merged = dict(old or {})
    if not new:
        return merged
    skip_empty = ("app_password", "appPassword")
    for k, v in new.items():
        if k in skip_empty and (v is None or str(v).strip() == ""):
            continue
        merged[k] = v
    return merged


def _merge_green_api_config(old: dict | None, new: dict | None) -> dict:
    merged = dict(old or {})
    if not new:
        return merged
    for k, v in new.items():
        if k in _SECRET_CFG_KEYS and (v is None or str(v).strip() == ""):
            continue
        merged[k] = v
    return merged


def _lead_read(lead: Lead) -> LeadRead:
    return LeadRead(
        id=lead.id,
        name=lead.name,
        phone=lead.phone,
        email=lead.email,
        source=lead.source,
        status_id=lead.status_id,
        stage_name=lead.stage.name if lead.stage else None,
        manager_id=lead.manager_id,
        refusal_reason=lead.refusal_reason,
        pipeline_id=lead.stage.pipeline_id if lead.stage else None,
        created_at=lead.created_at,
    )


async def _assert_pipeline_stage(db: AsyncSession, pipeline_id: int, stage_id: int, company_id: int) -> None:
    pipe = await db.get(Pipeline, pipeline_id)
    if pipe is None or pipe.company_id != company_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Pipeline not found")
    st = await db.get(PipelineStage, stage_id)
    if st is None or st.company_id != company_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Stage not found")
    if st.pipeline_id != pipeline_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Stage does not belong to pipeline")


def _provider_from_str(s: str) -> IntegrationProvider:
    try:
        return IntegrationProvider(s)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown provider")


@router.get("", response_model=list[IntegrationRead])
async def list_integrations(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    company_id: CurrentCompanyId,
) -> list[IntegrationRead]:
    await assert_owner_or_chief_expert(db, current_user)
    r = await db.execute(select(Integration).where(Integration.company_id == company_id).order_by(Integration.id.desc()))
    return [_integration_read(x) for x in r.scalars().all()]


@router.post("", response_model=IntegrationRead, status_code=status.HTTP_201_CREATED)
async def create_integration(
    body: IntegrationCreate,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    company_id: CurrentCompanyId,
) -> IntegrationRead:
    await assert_owner_or_chief_expert(db, current_user)
    provider = _provider_from_str(body.provider)
    await _assert_pipeline_stage(db, body.pipeline_id, body.stage_id, company_id)
    cfg = dict(body.config or {})
    sec = (body.secret or "").strip() if body.secret else ""
    if provider == IntegrationProvider.green_api:
        if not sec:
            sec = secrets.token_urlsafe(24)
        instance_id = cfg.get("instance_id") or cfg.get("instanceId")
        api_token = cfg.get("api_token") or cfg.get("apiToken") or cfg.get("apiTokenInstance")
        if not instance_id or not api_token:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Укажите idInstance и apiTokenInstance из личного кабинета Green API",
            )
    elif provider == IntegrationProvider.google_sheets:
        if not sec:
            sec = secrets.token_urlsafe(24)
        sheet_url = str(cfg.get("sheet_url") or cfg.get("spreadsheet_id") or "").strip()
        if not sheet_url:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Для Google Sheets укажите config.sheet_url (или spreadsheet_id)",
            )
    elif provider == IntegrationProvider.instagram:
        if len(sec) < 8:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Verify token (webhook-секрет) не короче 8 символов — сгенерируйте в форме или введите свой",
            )
        page_token = str(cfg.get("page_access_token") or cfg.get("pageAccessToken") or "").strip()
        if not page_token:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Укажите Page Access Token (Meta → инструменты Graph API или настройки Lead Ads)",
            )
    elif provider == IntegrationProvider.gmail:
        if not sec:
            sec = secrets.token_urlsafe(24)
        email = str(cfg.get("email") or cfg.get("gmail_email") or "").strip()
        app_password = str(cfg.get("app_password") or cfg.get("appPassword") or "").strip()
        if not email:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Для Gmail укажите email (или gmail_email) в config",
            )
        if not app_password:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Для Gmail укажите app_password (пароль приложения)",
            )
    else:
        if len(sec) < 8:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Webhook-секрет не короче 8 символов",
            )

    mx_int = await effective_tariff_max_integrations(db, company_id)
    if mx_int > 0:
        n_int = await count_company_integrations(db, company_id)
        if n_int >= mx_int:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Достигнут лимит интеграций по тарифу ({mx_int}).",
            )

    row = Integration(
        name=body.name.strip(),
        provider=provider,
        is_active=True,
        company_id=company_id,
        pipeline_id=body.pipeline_id,
        stage_id=body.stage_id,
        secret=sec,
        config=cfg,
        manager_close_deal_enabled=body.manager_close_deal_enabled,
    )
    db.add(row)
    await db.flush()
    await db.refresh(row)

    if provider == IntegrationProvider.green_api:
        instance_id = cfg.get("instance_id") or cfg.get("instanceId")
        api_token = cfg.get("api_token") or cfg.get("apiToken") or cfg.get("apiTokenInstance")
        pub = resolve_public_api_base(request, settings.public_api_base_url)
        if not pub:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Не удалось определить адрес API для автоподключения WhatsApp. "
                    "На сервере задайте public_api_base_url (например https://ваш-проект.amvera.io) "
                    "или откройте CRM с того же домена, что и бэкенд."
                ),
            )
        webhook_url = f"{pub}/api/integrations/webhook/{row.id}"
        api_base = green_api_base_from_config(cfg)
        ok, err = await asyncio.to_thread(
            push_green_incoming_webhook,
            instance_id=str(instance_id).strip(),
            api_token_instance=str(api_token).strip(),
            api_base=api_base,
            webhook_url=webhook_url,
            webhook_token=sec,
        )
        if not ok:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail=f"Green API не принял настройки (проверьте Instance ID, токен и адрес API в кабинете): {err}",
            )
        note = (
            "WhatsApp подключён: входящие сообщения будут попадать в выбранную воронку. "
            "Green API применяет настройки до 5 минут."
        )
        return _integration_read(row, setup_note=note)

    if provider == IntegrationProvider.instagram:
        pub = resolve_public_api_base(request, settings.public_api_base_url)
        hook = f"{pub}/api/integrations/webhook/{row.id}" if pub else f"/api/integrations/webhook/{row.id}"
        note = (
            f"Callback URL в Meta: {hook}. "
            "Verify token: тот же секрет, что в форме. Подписки: leadgen (страница), instagram, при необходимости messages. "
            "App Secret в config — для проверки подписи X-Hub-Signature-256 (рекомендуется)."
        )
        return _integration_read(row, setup_note=note)
    if provider == IntegrationProvider.gmail:
        note = (
            "Gmail подключён. Укажите email и app_password (пароль приложения Google). "
            "Для безопасного доступа включите IMAP в Gmail и используйте отдельный пароль приложения."
        )
        return _integration_read(row, setup_note=note)

    return _integration_read(row)


@router.patch("/{integration_id}", response_model=IntegrationRead)
async def patch_integration(
    integration_id: int,
    body: IntegrationUpdate,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    company_id: CurrentCompanyId,
) -> IntegrationRead:
    await assert_owner_or_chief_expert(db, current_user)

    row = await db.get(Integration, integration_id)
    if row is None or row.company_id != company_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration not found")

    if body.pipeline_id is not None or body.stage_id is not None:
        await _assert_pipeline_stage(db, body.pipeline_id or row.pipeline_id, body.stage_id or row.stage_id, company_id)
        if body.pipeline_id is not None:
            row.pipeline_id = body.pipeline_id
        if body.stage_id is not None:
            row.stage_id = body.stage_id

    if body.name is not None:
        row.name = body.name.strip()
    if body.is_active is not None:
        row.is_active = body.is_active
    if body.secret is not None:
        row.secret = body.secret.strip()
    if body.manager_close_deal_enabled is not None:
        row.manager_close_deal_enabled = body.manager_close_deal_enabled
    if body.config is not None:
        if row.provider == IntegrationProvider.green_api:
            merged = _merge_green_api_config(row.config, body.config)
            instance_id = merged.get("instance_id") or merged.get("instanceId")
            api_token = merged.get("api_token") or merged.get("apiToken") or merged.get("apiTokenInstance")
            if not instance_id or not api_token:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="For green_api set config.instance_id and config.api_token (или оставьте токен пустым при смене только instance)",
                )
            row.config = merged
        elif row.provider == IntegrationProvider.google_sheets:
            merged = dict(row.config or {})
            merged.update(body.config)
            sheet_url = str(merged.get("sheet_url") or merged.get("spreadsheet_id") or "").strip()
            if not sheet_url:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Для Google Sheets укажите config.sheet_url (или spreadsheet_id)",
                )
            row.config = merged
        elif row.provider == IntegrationProvider.instagram:
            merged = _merge_instagram_config(row.config, body.config)
            page_token = str(merged.get("page_access_token") or merged.get("pageAccessToken") or "").strip()
            if not page_token:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Нужен page_access_token в config (или оставьте прежний, не отправляя пустое поле)",
                )
            row.config = merged
        elif row.provider == IntegrationProvider.gmail:
            merged = _merge_gmail_config(row.config, body.config)
            email = str(merged.get("email") or merged.get("gmail_email") or "").strip()
            app_password = str(merged.get("app_password") or merged.get("appPassword") or "").strip()
            if not email:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Для Gmail укажите email (или gmail_email) в config",
                )
            if not app_password:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Для Gmail нужен app_password (или оставьте прежний, не отправляя пустое поле)",
                )
            row.config = merged
        else:
            row.config = body.config

    await db.flush()
    await db.refresh(row)

    if row.provider == IntegrationProvider.green_api:
        cfg = row.config or {}
        instance_id = cfg.get("instance_id") or cfg.get("instanceId")
        api_token = cfg.get("api_token") or cfg.get("apiToken") or cfg.get("apiTokenInstance")
        if instance_id and api_token:
            pub = resolve_public_api_base(request, settings.public_api_base_url)
            if not pub:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        "Не удалось определить адрес API. Задайте public_api_base_url на сервере "
                        "(например https://ваш-проект.amvera.io)."
                    ),
                )
            webhook_url = f"{pub}/api/integrations/webhook/{row.id}"
            api_base = green_api_base_from_config(cfg)
            ok, err = await asyncio.to_thread(
                push_green_incoming_webhook,
                instance_id=str(instance_id).strip(),
                api_token_instance=str(api_token).strip(),
                api_base=api_base,
                webhook_url=webhook_url,
                webhook_token=row.secret,
            )
            if not ok:
                raise HTTPException(
                    status_code=status.HTTP_502_BAD_GATEWAY,
                    detail=f"Green API не принял настройки: {err}",
                )
            return _integration_read(
                row,
                setup_note="Настройки WhatsApp отправлены в Green API (применение до 5 минут).",
            )

    return _integration_read(row)


class GreenWebhookStatusRead(BaseModel):
    integration_id: int
    expected_webhook_url: str
    public_api_base_url: str
    green_webhook_url: str | None = None
    green_incoming_webhook: str | None = None
    green_state_instance: str | None = None
    webhook_url_matches: bool = False
    incoming_enabled: bool = False
    instance_authorized: bool | None = None
    sync_error: str | None = None
    hint: str | None = None


async def _green_webhook_apply(
    request: Request,
    row: Integration,
) -> tuple[str, str]:
    cfg = row.config or {}
    instance_id = cfg.get("instance_id") or cfg.get("instanceId")
    api_token = cfg.get("api_token") or cfg.get("apiToken") or cfg.get("apiTokenInstance")
    if not instance_id or not api_token:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="У интеграции нет instance_id или api_token в config",
        )
    pub = resolve_public_api_base(request, settings.public_api_base_url)
    if not pub:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Задайте PUBLIC_API_BASE_URL на сервере Amvera "
                "(например https://metodi-one-koujikin.amvera.io) и пересохраните интеграцию."
            ),
        )
    webhook_url = f"{pub.rstrip('/')}/api/integrations/webhook/{row.id}"
    api_base = green_api_base_from_config(cfg)
    ok, err = await asyncio.to_thread(
        push_green_incoming_webhook,
        instance_id=str(instance_id).strip(),
        api_token_instance=str(api_token).strip(),
        api_base=api_base,
        webhook_url=webhook_url,
        webhook_token=row.secret,
    )
    if not ok:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Green API не принял настройки: {err}",
        )
    return webhook_url, pub


def _green_webhook_status_hint(
    *,
    expected: str,
    actual: str | None,
    incoming: str | None,
    authorized: bool | None,
) -> str | None:
    if authorized is False:
        return "Инстанс WhatsApp не авторизован — отсканируйте QR в кабинете Green API."
    if not actual:
        return "В Green API не задан webhookUrl. Нажмите «Переподключить webhook» в CRM."
    if actual.rstrip("/") != expected.rstrip("/"):
        return (
            "Webhook в Green API указывает на другой адрес. "
            "Нажмите «Переподключить webhook» — CRM пропишет правильный URL на Amvera."
        )
    if (incoming or "").lower() != "yes":
        return "В Green API выключен incomingWebhook. Переподключите webhook из CRM."
    return None


@router.get("/{integration_id}/green-webhook-status", response_model=GreenWebhookStatusRead)
async def green_webhook_status(
    integration_id: int,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    company_id: CurrentCompanyId,
) -> GreenWebhookStatusRead:
    if current_user.role != UserRole.owner:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Только владелец")
    row = await db.get(Integration, integration_id)
    if row is None or row.company_id != company_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration not found")
    if row.provider != IntegrationProvider.green_api:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Только для Green API")

    pub = resolve_public_api_base(request, settings.public_api_base_url)
    expected = f"{pub.rstrip('/')}/api/integrations/webhook/{row.id}" if pub else ""
    cfg = row.config or {}
    instance_id = str(cfg.get("instance_id") or cfg.get("instanceId") or "").strip()
    api_token = str(cfg.get("api_token") or cfg.get("apiToken") or cfg.get("apiTokenInstance") or "").strip()
    api_base = green_api_base_from_config(cfg)

    out = GreenWebhookStatusRead(
        integration_id=row.id,
        expected_webhook_url=expected,
        public_api_base_url=pub or settings.public_api_base_url or "",
    )
    if not pub:
        out.hint = (
            "На Amvera не задан PUBLIC_API_BASE_URL. "
            "Добавьте переменную и нажмите «Переподключить webhook»."
        )
        return out
    if not instance_id or not api_token:
        out.hint = "Укажите Instance ID и API Token в настройках интеграции."
        return out

    ok, data = await asyncio.to_thread(
        fetch_green_settings,
        instance_id=instance_id,
        api_token_instance=api_token,
        api_base=api_base,
    )
    if not ok:
        out.sync_error = str(data)
        out.hint = "Не удалось прочитать настройки Green API — проверьте Instance ID, токен и api_base_url."
        return out

    settings_body = data
    if isinstance(settings_body, dict) and isinstance(settings_body.get("settings"), dict):
        settings_body = settings_body["settings"]

    green_url = str((settings_body or {}).get("webhookUrl") or "").strip() or None
    incoming = str((settings_body or {}).get("incomingWebhook") or "").strip() or None
    green_token = _normalize_webhook_token(str((settings_body or {}).get("webhookUrlToken") or "").strip() or None)
    crm_token = _normalize_webhook_token(row.secret)
    token_matches = bool(green_token and crm_token and green_token == crm_token)

    state_ok, state_val = await asyncio.to_thread(
        fetch_green_state_instance,
        instance_id=instance_id,
        api_token_instance=api_token,
        api_base=api_base,
    )
    state = str(state_val).strip() if state_ok and isinstance(state_val, str) else None

    out.green_webhook_url = green_url
    out.green_incoming_webhook = incoming
    out.green_state_instance = state
    out.webhook_url_matches = bool(green_url and green_url.rstrip("/") == expected.rstrip("/"))
    out.incoming_enabled = (incoming or "").lower() == "yes"
    if state:
        out.instance_authorized = state.lower() in ("authorized", "starting")
    out.hint = _green_webhook_status_hint(
        expected=expected,
        actual=green_url,
        incoming=incoming,
        authorized=out.instance_authorized,
    )
    if out.hint is None and green_token and crm_token and not token_matches:
        out.hint = (
            "Токен webhook в Green API не совпадает с секретом интеграции в CRM. "
            "Нажмите «Переподключить webhook»."
        )
    return out


@router.post("/{integration_id}/green-webhook-sync", response_model=IntegrationRead)
async def green_webhook_sync(
    integration_id: int,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    company_id: CurrentCompanyId,
) -> IntegrationRead:
    if current_user.role != UserRole.owner:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Только владелец")
    row = await db.get(Integration, integration_id)
    if row is None or row.company_id != company_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration not found")
    if row.provider != IntegrationProvider.green_api:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Только для Green API")
    if not row.is_active:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Интеграция выключена")

    await _green_webhook_apply(request, row)
    return _integration_read(
        row,
        setup_note=(
            "Webhook WhatsApp отправлен в Green API. Применение до 5 минут. "
            "После этого входящие сообщения появятся в разделе «Чаты»."
        ),
    )


@router.post("/{integration_id}/sync")
async def sync_integration_now(
    integration_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    company_id: CurrentCompanyId,
) -> dict[str, int]:
    await assert_owner_or_chief_expert(db, current_user)
    row = await db.get(Integration, integration_id)
    if row is None or row.company_id != company_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration not found")
    if row.provider != IntegrationProvider.google_sheets:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Sync supported only for Google Sheets")
    try:
        stats = await sync_google_sheet_integration(db, integ=row, max_rows=1000)
    except RuntimeError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    await db.commit()
    return stats


class GreenBackfillResult(BaseModel):
    days: int
    chats_scanned: int
    chats_imported: int
    leads_created: int
    leads_updated: int
    messages_added: int
    skipped_answered: int
    skipped_no_match: int
    errors: list[str] = Field(default_factory=list)


@router.post("/{integration_id}/green-backfill", response_model=GreenBackfillResult)
async def green_api_backfill_now(
    integration_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    company_id: CurrentCompanyId,
    days: int = Query(default=7, ge=1, le=30),
) -> GreenBackfillResult:
    """
    Догрузка пропущенных WhatsApp-диалогов из журнала Green API.
    Импортирует чаты за последние N дней, где клиент написал после приветствия
    (или при сбое webhook — написал вообще) и не получил ответ менеджера.
    """
    await assert_owner_or_chief_expert(db, current_user)
    row = await db.get(Integration, integration_id)
    if row is None or row.company_id != company_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration not found")
    if row.provider != IntegrationProvider.green_api:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Догрузка доступна только для Green API")
    if not row.is_active:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Интеграция выключена")
    try:
        stats = await sync_green_api_backfill(db, integ=row, company_id=company_id, days=days)
    except RuntimeError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    await db.commit()
    return GreenBackfillResult(**stats)


class GreenBroadcastResult(BaseModel):
    requested_count: int
    sent_count: int
    failed_count: int
    failed_numbers: list[str] = Field(default_factory=list)


class GreenBroadcastPreviewRead(BaseModel):
    found_count: int
    unique_count: int
    limited_count: int


async def _collect_green_broadcast_phones(
    *,
    db: AsyncSession,
    company_id: int,
    source: str,
    excel_phone_column: str,
    file: UploadFile | None,
) -> list[str]:
    src = source.strip().lower()
    phones: list[str] = []
    if src == "database":
        phone_rows = (
            await db.execute(
                select(Lead.phone)
                .where(Lead.company_id == company_id, Lead.phone.is_not(None))
                .order_by(Lead.id.desc())
                .limit(5000)
            )
        ).all()
        phones = [_norm_phone(str(x[0] or "")) for x in phone_rows]
        phones = [p for p in phones if p]
    elif src == "excel":
        if file is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="При source=excel прикрепите файл")
        content = await file.read()
        if not content:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Файл пуст")
        filename = (file.filename or "").lower()
        if filename.endswith(".xlsx"):
            phones = _extract_phones_from_excel(content, phone_column=excel_phone_column)
        elif filename.endswith(".csv") or not filename:
            phones = _extract_phones_from_csv(content)
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Поддерживаются только .xlsx и .csv")
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="source должен быть database или excel")
    seen: set[str] = set()
    uniq: list[str] = []
    for p in phones:
        if p and p not in seen:
            seen.add(p)
            uniq.append(p)
    return uniq


@router.post("/{integration_id}/green-broadcast/preview", response_model=GreenBroadcastPreviewRead)
async def green_broadcast_preview(
    integration_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    company_id: CurrentCompanyId,
    source: Annotated[str, Form()] = "database",
    excel_phone_column: Annotated[str, Form()] = "phone",
    file: UploadFile | None = File(default=None),
) -> GreenBroadcastPreviewRead:
    await assert_owner_or_chief_expert(db, current_user)
    row = await db.get(Integration, integration_id)
    if row is None or row.company_id != company_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration not found")
    if row.provider != IntegrationProvider.green_api:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Предпросмотр доступен только для Green API")
    phones = await _collect_green_broadcast_phones(
        db=db,
        company_id=company_id,
        source=source,
        excel_phone_column=excel_phone_column,
        file=file,
    )
    return GreenBroadcastPreviewRead(
        found_count=len(phones),
        unique_count=len(phones),
        limited_count=min(len(phones), 2000),
    )


@router.post("/{integration_id}/green-broadcast", response_model=GreenBroadcastResult)
async def green_broadcast(
    integration_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    company_id: CurrentCompanyId,
    message: Annotated[str, Form(min_length=1, max_length=4096)],
    source: Annotated[str, Form()] = "database",
    excel_phone_column: Annotated[str, Form()] = "phone",
    file: UploadFile | None = File(default=None),
) -> GreenBroadcastResult:
    await assert_owner_or_chief_expert(db, current_user)
    row = await db.get(Integration, integration_id)
    if row is None or row.company_id != company_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration not found")
    if row.provider != IntegrationProvider.green_api:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Рассылка доступна только для Green API")
    if not row.is_active:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Интеграция выключена")
    phones = await _collect_green_broadcast_phones(
        db=db,
        company_id=company_id,
        source=source,
        excel_phone_column=excel_phone_column,
        file=file,
    )
    phones = phones[:2000]
    if not phones:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Не найдено валидных номеров для рассылки")

    text = message.strip()
    sent = 0
    failed: list[str] = []
    for phone in phones:
        ok, err, _ = await send_green_text_async(row.config, _to_green_chat_id(phone), text)
        if ok:
            sent += 1
        else:
            logger.warning("green broadcast failed integration_id=%s phone=%s err=%s", integration_id, phone, err)
            failed.append(phone)
    return GreenBroadcastResult(
        requested_count=len(phones),
        sent_count=sent,
        failed_count=len(failed),
        failed_numbers=failed[:100],
    )


@router.post("/generate-secret")
async def generate_secret(
    current_user: CurrentUser,
) -> dict[str, str]:
    await assert_owner_or_chief_expert(db, current_user)
    return {"secret": secrets.token_urlsafe(24)}


def _to_green_chat_id(phone: str) -> str:
    return f"{phone}@c.us"


async def _create_lead_from_integration(
    db: AsyncSession,
    *,
    integ: Integration,
    company_id: int,
    name: str,
    phone: str | None,
    email: str | None,
    source_name: str,
    external_chat_id: str | None = None,
    thread_provider: str | None = None,
) -> Lead:
    lead, _ = await create_lead_from_integration(
        db,
        integ=integ,
        company_id=company_id,
        name=name,
        phone=phone,
        email=email,
        source_name=source_name,
        external_chat_id=external_chat_id,
        thread_provider=thread_provider,
    )
    return lead


def _extract_phones_from_excel(content: bytes, *, phone_column: str) -> list[str]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    first = next(rows, None)
    if not first:
        return []
    header = [str(x).strip().lower() if x is not None else "" for x in first]
    wanted = phone_column.strip().lower()
    idx = header.index(wanted) if wanted in header else 0
    out: list[str] = []
    for r in rows:
        if r is None or idx >= len(r):
            continue
        p = _norm_phone(str(r[idx] or ""))
        if p:
            out.append(p)
    return out


def _extract_phones_from_csv(content: bytes) -> list[str]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    out: list[str] = []
    for row in reader:
        if not row:
            continue
        p = _norm_phone(row[0])
        if p:
            out.append(p)
    return out


@router.get("/webhook/{integration_id}")
async def integration_webhook_verify(
    integration_id: int,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Meta при подписке webhook шлёт GET с hub.mode / hub.verify_token / hub.challenge."""
    integ = await db.get(Integration, integration_id)
    if integ is None or not integ.is_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration not found")
    if integ.provider != IntegrationProvider.instagram:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Use this URL only for Instagram / Meta")
    qp = request.query_params
    challenge = meta_hub_challenge_response(
        verify_token=integ.secret,
        hub_mode=qp.get("hub.mode"),
        hub_verify_token=qp.get("hub.verify_token"),
        hub_challenge=qp.get("hub.challenge"),
    )
    if challenge is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Verification failed")
    return challenge


@router.post("/webhook/{integration_id}", response_model=None)
async def integration_webhook(
    integration_id: int,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    token: str | None = Query(default=None),
    x_webhook_token: str | None = Header(default=None),
    authorization: Annotated[str | None, Header()] = None,
) -> Any:
    raw_body = await request.body()
    integ = await db.get(Integration, integration_id)
    if integ is None or not integ.is_active:
        logger.warning("integration webhook: integration_id=%s not found or inactive", integration_id)
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration not found")
    if integ.company_id is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Integration has no company scope")
    company_id = int(integ.company_id)

    if integ.provider == IntegrationProvider.instagram:
        sig = request.headers.get("X-Hub-Signature-256")
        try:
            payload: Any = json.loads(raw_body.decode("utf-8") or "{}") if raw_body else {}
        except Exception:
            logger.exception("integration webhook: integration_id=%s invalid JSON body (instagram)", integration_id)
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Expected JSON body")
        if not isinstance(payload, dict):
            payload = {}
        return await handle_instagram_webhook(
            db,
            integ=integ,
            company_id=company_id,
            raw_body=raw_body,
            payload=payload,
            signature_header=sig,
            create_lead_fn=_create_lead_from_integration,
            upsert_thread_fn=_upsert_thread,
            add_message_fn=_add_incoming_message,
            lead_read_fn=_lead_read,
        )

    provided = token or x_webhook_token or _token_from_authorization_header(authorization)
    if not _webhook_token_matches(provided, integ.secret):
        logger.warning(
            "integration webhook: integration_id=%s rejected (missing/wrong token, len_provided=%s)",
            integration_id,
            len(provided or ""),
        )
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bad token")

    try:
        payload = json.loads(raw_body.decode("utf-8") or "{}") if raw_body else {}
    except Exception:
        logger.exception("integration webhook: integration_id=%s invalid JSON body", integration_id)
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Expected JSON body")

    logger.info(
        "integration webhook: integration_id=%s provider=%s typeWebhook=%s",
        integration_id,
        integ.provider.value if hasattr(integ.provider, "value") else integ.provider,
        payload.get("typeWebhook") if isinstance(payload, dict) else None,
    )

    # MVP парсеры под провайдеры
    if integ.provider == IntegrationProvider.telegram:
        msg = payload.get("message") or payload.get("edited_message") or {}
        frm = msg.get("from") or {}
        chat = msg.get("chat") or {}
        text = msg.get("text") or ""
        name = (frm.get("first_name") or "") + (" " + (frm.get("last_name") or "") if frm.get("last_name") else "")
        name = name.strip() or (chat.get("title") or "Telegram lead")
        username = frm.get("username")
        source_name = "TELEGRAM"
        phone = None
        contact = msg.get("contact") or {}
        if isinstance(contact, dict):
            phone = contact.get("phone_number")
        if username:
            text = f"@{username}: {text}".strip()
        ext_chat = str(chat.get("id") or "").strip()
        lead = await _create_lead_from_integration(
            db,
            integ=integ,
            company_id=company_id,
            name=name,
            phone=phone,
            email=None,
            source_name=source_name,
            external_chat_id=ext_chat or None,
            thread_provider=IntegrationProvider.telegram.value,
        )
        thread = await _upsert_thread(
            db,
            company_id=company_id,
            lead=lead,
            provider=IntegrationProvider.telegram.value,
            external_chat_id=ext_chat or None,
            title=name,
        )
        await _add_incoming_message(db, company_id, thread.id, text)
        logger.info("integration webhook: ok lead_id=%s thread_id=%s", lead.id, thread.id)
        return _lead_read(lead)

    if integ.provider == IntegrationProvider.green_api:
        # Иначе на каждый stateInstance/outgoingStatus и т.д. создавались бы новые лиды
        tw = str(payload.get("typeWebhook") or "").strip()
        # Green API обычно шлёт `incomingMessageReceived`; регистр в редких случаях может отличаться.
        if tw.lower() != "incomingmessagereceived":
            logger.info(
                "integration webhook: integration_id=%s green_api skip typeWebhook=%s",
                integration_id,
                tw or "(empty)",
            )
            return Response(status_code=204)
        # Ожидаем типичный webhook GREEN API: { "typeWebhook": "...", "senderData": {...}, "messageData": {...} }
        sender = payload.get("senderData") or {}
        message_data = payload.get("messageData") or {}
        raw_chat = sender.get("chatId") or payload.get("chatId")
        chat_id = str(raw_chat).strip() if raw_chat not in (None, "") else ""
        sender_name = sender.get("senderName") or "WhatsApp lead"
        phone = None
        if chat_id.endswith("@c.us"):
            phone = chat_id.replace("@c.us", "")
        elif chat_id:
            phone = chat_id
        # Иногда номер в sender
        if not phone:
            phone = sender.get("sender") or sender.get("senderPhone")
        # Имя из сообщения может быть полезнее
        if not sender_name:
            sender_name = (
                message_data.get("fileMessageData", {}).get("caption")
                or message_data.get("extendedTextMessageData", {}).get("text")
                or "WhatsApp lead"
            )
        source_name = "GREEN API"
        text, mtype, murl, mmime, mfn = parse_green_message_data(message_data)
        green_msg_id = str(payload.get("idMessage") or "").strip() or None
        ext_chat = chat_id or None
        lead = await _create_lead_from_integration(
            db,
            integ=integ,
            company_id=company_id,
            name=str(sender_name),
            phone=phone,
            email=None,
            source_name=source_name,
            external_chat_id=ext_chat,
            thread_provider=IntegrationProvider.green_api.value,
        )
        thread = await _upsert_thread(
            db,
            company_id=company_id,
            lead=lead,
            provider=IntegrationProvider.green_api.value,
            external_chat_id=ext_chat,
            title=str(sender_name),
            pipeline_id=int(integ.pipeline_id),
        )
        incoming_msg = await _add_incoming_message(
            db,
            company_id,
            thread.id,
            text,
            message_type=mtype,
            media_url=murl,
            media_mime=mmime,
            file_name=mfn,
            provider_message_id=green_msg_id,
        )
        if incoming_msg is None and green_msg_id:
            incoming_msg = await _find_incoming_message_by_provider_id(db, company_id, green_msg_id)
        await persist_incoming_green_media_if_needed(
            db,
            msg=incoming_msg,
            config=integ.config if isinstance(integ.config, dict) else {},
            chat_id=chat_id,
            id_message=green_msg_id,
            message_type=mtype,
            download_url=murl,
            file_name=mfn,
            media_mime=mmime,
        )
        await send_welcome_if_first_incoming(
            db,
            lead=lead,
            thread=thread,
            integration=integ,
        )
        logger.info("integration webhook: ok lead_id=%s thread_id=%s", lead.id, thread.id)
        return _lead_read(lead)

    if integ.provider == IntegrationProvider.google_sheets:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Google Sheets: используйте синхронизацию из панели интеграций, webhook не применяется",
        )
    if integ.provider == IntegrationProvider.gmail:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Gmail: webhook не используется, подключение настраивается через config в панели интеграций",
        )

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=f"Webhook для провайдера не поддерживается: {integ.provider}",
    )

