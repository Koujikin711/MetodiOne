"""Экспорт ключевых финотчётов в один XLSX (несколько листов)."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.finance_reports import (
    balance_sheet_snapshot,
    cash_flow_statement,
    pl_by_account,
    trial_balance_rows,
)


async def build_finance_workbook_bytes(
    db: AsyncSession,
    company_id: int,
    date_from: datetime,
    date_to: datetime,
) -> bytes:
    wb = Workbook()
    default = wb.active
    assert default is not None
    default.title = "ОСВ"

    tb = await trial_balance_rows(db, company_id, date_from, date_to)
    default.append(["Код", "Название", "Тип", "Дебет", "Кредит", "Сальдо Дт−Кт"])
    for code, name, atype, d, c, net in tb:
        default.append([code, name, atype, float(d), float(c), float(net)])

    ws_pl = wb.create_sheet("ОПиУ")
    pl_rows = await pl_by_account(db, company_id, date_from, date_to, ("revenue", "expense"))
    ws_pl.append(["Код", "Название", "Тип", "Сумма (период)"])
    for code, name, atype, amt in pl_rows:
        ws_pl.append([code, name, atype, float(amt)])

    ws_bs = wb.create_sheet("Баланс")
    _d0, d1 = date_from, date_to
    asset_lines, liability_lines, equity_lines, ta, tld, tea, ret, tp, bal = await balance_sheet_snapshot(db, company_id, d1)
    ws_bs.append(["Раздел", "Код", "Название", "Сумма"])
    ws_bs.append(["asset", "", "Актив (итог)", float(ta)])
    for code, name, amt in asset_lines:
        ws_bs.append(["asset", code, name, float(amt)])
    ws_bs.append(["liability", "", "Обязательства (итог)", float(tld)])
    for code, name, amt in liability_lines:
        ws_bs.append(["liability", code, name, float(amt)])
    ws_bs.append(["equity", "", "Капитал (счета)", float(tea)])
    for code, name, amt in equity_lines:
        ws_bs.append(["equity", code, name, float(amt)])
    ws_bs.append(["retained", "", "Накопленная прибыль (ОПиУ)", float(ret)])
    ws_bs.append(["total", "", "Пассив (итог)", float(tp)])
    ws_bs.append(["check", "", "Баланс сходится", bool(bal)])

    ws_cf = wb.create_sheet("ДДС")
    opening, closing, net_ch, buckets = await cash_flow_statement(db, company_id, date_from, date_to)
    ws_cf.append(["Показатель", "Значение"])
    ws_cf.append(["Остаток ДС на начало", float(opening)])
    ws_cf.append(["Остаток ДС на конец", float(closing)])
    ws_cf.append(["Чистое изменение", float(net_ch)])
    ws_cf.append([])
    ws_cf.append(["Статья (ключ)", "Подпись", "Сумма"])
    for key, label, amt in buckets:
        ws_cf.append([key, label, float(amt)])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
