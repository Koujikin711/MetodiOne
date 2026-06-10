"""Интеграция ОСВ: Gmail (IMAP) + данные CRM."""

from __future__ import annotations

import asyncio
import csv
import email
import imaplib
import io
import re
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from email.header import decode_header
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    BookingAppointment,
    Deal,
    FinanceCompanySettings,
    FinanceOsvRow,
    Integration,
    IntegrationProvider,
    Lead,
)

_HEADER_ALIASES: dict[str, str] = {
    "дата": "txn_date",
    "date": "txn_date",
    "маблаги партном": "partner_amount",
    "период оказания услуги": "service_period",
    "выручка": "revenue",
    "выручка - som": "revenue",
    "revenue": "revenue",
    "расход": "expense",
    "расход - som": "expense",
    "expense": "expense",
    "банк": "bank",
    "основание выручка/расход": "basis",
    "основание": "basis",
    "контрагенты": "counterparty",
    "контрагент": "counterparty",
    "телефон": "phone",
    "чрз": "via_person",
    "товар/услуга": "product_service",
    "товар": "product_service",
    "услуга": "product_service",
    "статья": "article",
    "подробно": "detail_category",
    "кратко": "brief_category",
}

_AMOUNT_RE = re.compile(r"(\d[\d\s.,]{2,})\s*(?:som|сом|tjs|₽|руб)?", re.I)


def _decode_mime_header(raw: str | None) -> str:
    if not raw:
        return ""
    parts: list[str] = []
    for chunk, enc in decode_header(raw):
        if isinstance(chunk, bytes):
            parts.append(chunk.decode(enc or "utf-8", errors="replace"))
        else:
            parts.append(str(chunk))
    return "".join(parts)


def _parse_decimal(raw: Any) -> Decimal:
    if raw is None:
        return Decimal("0")
    if isinstance(raw, (int, float, Decimal)):
        return Decimal(str(raw))
    s = str(raw).strip().replace("\u00a0", "").replace(" ", "")
    if not s or s in ("-", "—"):
        return Decimal("0")
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _parse_date(raw: Any) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def _normalize_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def _row_from_mapping(data: dict[str, Any]) -> dict[str, Any] | None:
    txn_date = _parse_date(data.get("txn_date"))
    revenue = _parse_decimal(data.get("revenue"))
    expense = _parse_decimal(data.get("expense"))
    if txn_date is None or (revenue <= 0 and expense <= 0):
        return None
    return {
        "txn_date": txn_date,
        "partner_amount": _parse_decimal(data.get("partner_amount")) or None,
        "service_period": (data.get("service_period") or None),
        "revenue": revenue,
        "expense": expense,
        "bank": (str(data.get("bank") or "").strip() or None),
        "basis": (str(data.get("basis") or "").strip() or None),
        "counterparty": (str(data.get("counterparty") or "").strip() or None),
        "phone": (str(data.get("phone") or "").strip() or None),
        "via_person": (str(data.get("via_person") or "").strip() or None),
        "product_service": (str(data.get("product_service") or "").strip() or None),
        "article": (str(data.get("article") or "").strip() or None),
        "detail_category": (str(data.get("detail_category") or "").strip() or None),
        "brief_category": (str(data.get("brief_category") or "").strip() or None),
    }


def _parse_csv_bytes(raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if len(rows) < 2:
        return []
    header = [_normalize_header(c) for c in rows[0]]
    col_map: dict[int, str] = {}
    for i, h in enumerate(header):
        field = _HEADER_ALIASES.get(h)
        if field:
            col_map[i] = field
    if not col_map:
        return []
    out: list[dict[str, Any]] = []
    for line in rows[1:]:
        data: dict[str, Any] = {}
        for i, field in col_map.items():
            if i < len(line):
                data[field] = line[i]
        parsed = _row_from_mapping(data)
        if parsed:
            out.append(parsed)
    return out


def _parse_xlsx_bytes(raw: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = [_normalize_header(str(c or "")) for c in rows[0]]
    col_map: dict[int, str] = {}
    for i, h in enumerate(header):
        field = _HEADER_ALIASES.get(h)
        if field:
            col_map[i] = field
    out: list[dict[str, Any]] = []
    for line in rows[1:]:
        data: dict[str, Any] = {}
        for i, field in col_map.items():
            if i < len(line):
                data[field] = line[i]
        parsed = _row_from_mapping(data)
        if parsed:
            out.append(parsed)
    return out


def _parse_email_body_row(subject: str, body: str, fallback_date: date) -> dict[str, Any] | None:
    text = f"{subject}\n{body}"
    amounts = _AMOUNT_RE.findall(text)
    if not amounts:
        return None
    amt = _parse_decimal(amounts[0])
    if amt <= 0:
        return None
    low = text.lower()
    is_expense = any(w in low for w in ("списан", "расход", "оплат", "перевод", "снят"))
    revenue = Decimal("0") if is_expense else amt
    expense = amt if is_expense else Decimal("0")
    brief = "Расход" if is_expense else "Выручка"
    article = "Поступления" if not is_expense else "Административные расходы"
    return {
        "txn_date": fallback_date,
        "revenue": revenue,
        "expense": expense,
        "basis": subject[:255] if subject else None,
        "brief_category": brief,
        "article": article,
        "bank": "Gmail",
    }


def _imap_fetch_parsed_rows(email_addr: str, app_password: str, imap_host: str) -> list[tuple[str, dict[str, Any]]]:
    out: list[tuple[str, dict[str, Any]]] = []
    mail = imaplib.IMAP4_SSL(imap_host or "imap.gmail.com")
    try:
        mail.login(email_addr, app_password)
        mail.select("INBOX")
        _status, data = mail.search(None, "ALL")
        ids = (data[0] or b"").split()
        for msg_id in ids[-80:]:
            _status, msg_data = mail.fetch(msg_id, "(RFC822)")
            if not msg_data or not msg_data[0]:
                continue
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)
            subject = _decode_mime_header(msg.get("Subject"))
            ext_base = f"gmail:{msg_id.decode() if isinstance(msg_id, bytes) else msg_id}"
            msg_date = email.utils.parsedate_to_datetime(msg.get("Date") or "")
            fallback_date = msg_date.date() if msg_date else date.today()
            body_text = ""
            if msg.is_multipart():
                for part in msg.walk():
                    ctype = part.get_content_type()
                    fname = part.get_filename()
                    payload = part.get_payload(decode=True) or b""
                    if fname:
                        low = fname.lower()
                        parsed_rows: list[dict[str, Any]] = []
                        if low.endswith(".csv"):
                            parsed_rows = _parse_csv_bytes(payload)
                        elif low.endswith((".xlsx", ".xlsm")):
                            parsed_rows = _parse_xlsx_bytes(payload)
                        for i, row in enumerate(parsed_rows):
                            out.append((f"{ext_base}:att:{fname}:{i}", row))
                    elif ctype == "text/plain" and not body_text:
                        body_text = payload.decode("utf-8", errors="replace")
            else:
                body_text = (msg.get_payload(decode=True) or b"").decode("utf-8", errors="replace")
            if body_text:
                single = _parse_email_body_row(subject, body_text, fallback_date)
                if single:
                    out.append((f"{ext_base}:body", single))
    finally:
        try:
            mail.logout()
        except Exception:
            pass
    return out


async def _get_gmail_integration(db: AsyncSession, company_id: int) -> Integration | None:
    row = (
        await db.execute(
            select(Integration)
            .where(
                Integration.company_id == company_id,
                Integration.provider == IntegrationProvider.gmail,
                Integration.is_active.is_(True),
            )
            .limit(1),
        )
    ).scalars().first()
    return row


async def _upsert_osv_row(
    db: AsyncSession,
    *,
    company_id: int,
    external_key: str,
    source: str,
    payload: dict[str, Any],
) -> bool:
    exists = (
        await db.execute(
            select(FinanceOsvRow.id).where(
                FinanceOsvRow.company_id == company_id,
                FinanceOsvRow.external_key == external_key,
            ).limit(1),
        )
    ).scalar_one_or_none()
    if exists is not None:
        return False
    db.add(
        FinanceOsvRow(
            company_id=company_id,
            source=source,
            external_key=external_key,
            **payload,
        ),
    )
    return True


async def sync_crm_to_osv(db: AsyncSession, company_id: int) -> tuple[int, int]:
    imported = 0
    skipped = 0
    deals = (
        await db.execute(
            select(Deal, Lead.name, Lead.phone)
            .join(Lead, Lead.id == Deal.lead_id, isouter=True)
            .where(
                Deal.paid_amount > 0,
                (Deal.company_id == company_id) | (Lead.company_id == company_id),
            ),
        )
    ).all()
    for deal, lead_name, lead_phone in deals:
        paid = Decimal(deal.paid_amount or 0)
        if paid <= 0:
            continue
        payload = {
            "txn_date": date.today(),
            "revenue": paid,
            "expense": Decimal("0"),
            "counterparty": lead_name,
            "phone": lead_phone,
            "product_service": deal.title or "Закрытая сделка",
            "article": "Поступления",
            "brief_category": "Выручка",
            "basis": f"CRM сделка #{deal.id}",
            "bank": "CRM",
        }
        ok = await _upsert_osv_row(
            db,
            company_id=company_id,
            external_key=f"crm:deal:{deal.id}",
            source="crm",
            payload=payload,
        )
        if ok:
            imported += 1
        else:
            skipped += 1

    appts = (
        await db.execute(
            select(BookingAppointment, Lead.name, Lead.phone)
            .join(Lead, Lead.id == BookingAppointment.lead_id, isouter=True)
            .where(
                BookingAppointment.company_id == company_id,
                BookingAppointment.paid_amount > 0,
            ),
        )
    ).all()
    for appt, lead_name, lead_phone in appts:
        paid = Decimal(appt.paid_amount or 0)
        if paid <= 0:
            continue
        payload = {
            "txn_date": (appt.start_at or datetime.now(UTC)).date(),
            "revenue": paid,
            "expense": Decimal("0"),
            "counterparty": lead_name,
            "phone": lead_phone,
            "product_service": appt.service_title or "Онлайн-запись",
            "article": "Поступления",
            "brief_category": "Выручка",
            "basis": f"Оплата записи #{appt.id}",
            "bank": "CRM",
        }
        ok = await _upsert_osv_row(
            db,
            company_id=company_id,
            external_key=f"crm:appt:{appt.id}",
            source="crm",
            payload=payload,
        )
        if ok:
            imported += 1
        else:
            skipped += 1
    return imported, skipped


async def sync_gmail_to_osv(db: AsyncSession, company_id: int) -> tuple[int, int, bool, str | None]:
    integ = await _get_gmail_integration(db, company_id)
    if integ is None:
        return 0, 0, False, None
    cfg = integ.config if isinstance(integ.config, dict) else {}
    email_addr = str(cfg.get("email") or cfg.get("gmail_email") or "").strip()
    app_password = str(cfg.get("app_password") or "").strip()
    imap_host = str(cfg.get("imap_host") or "imap.gmail.com").strip()
    if not email_addr or not app_password:
        return 0, 0, False, email_addr or None

    parsed = await asyncio.to_thread(_imap_fetch_parsed_rows, email_addr, app_password, imap_host)
    imported = 0
    skipped = 0
    for external_key, payload in parsed:
        ok = await _upsert_osv_row(
            db,
            company_id=company_id,
            external_key=external_key,
            source="gmail",
            payload=payload,
        )
        if ok:
            imported += 1
        else:
            skipped += 1
    return imported, skipped, True, email_addr


async def run_finance_integrate(db: AsyncSession, company_id: int) -> dict[str, Any]:
    gmail_imported, gmail_skipped, gmail_ok, gmail_email = await sync_gmail_to_osv(db, company_id)
    crm_imported, crm_skipped = await sync_crm_to_osv(db, company_id)

    settings = (
        await db.execute(select(FinanceCompanySettings).where(FinanceCompanySettings.company_id == company_id))
    ).scalars().first()
    if settings is None:
        settings = FinanceCompanySettings(company_id=company_id)
        db.add(settings)
    settings.updated_at = datetime.now(UTC)
    if gmail_imported or crm_imported:
        today = date.today()
        settings.last_osv_import_to = today
        if settings.last_osv_import_from is None:
            settings.last_osv_import_from = date(today.year, 1, 1)

    total = (
        await db.execute(
            select(FinanceOsvRow.id).where(FinanceOsvRow.company_id == company_id),
        )
    ).all()

    msg_parts: list[str] = []
    if gmail_ok:
        msg_parts.append(f"Gmail: +{gmail_imported} строк")
    else:
        msg_parts.append("Gmail не подключён — подключите в «Интеграции»")
    msg_parts.append(f"CRM: +{crm_imported} строк")
    skipped = gmail_skipped + crm_skipped
    if skipped:
        msg_parts.append(f"пропущено дублей: {skipped}")

    return {
        "gmail_connected": gmail_ok,
        "gmail_email": gmail_email,
        "imported_from_gmail": gmail_imported,
        "imported_from_crm": crm_imported,
        "skipped_duplicates": skipped,
        "osv_rows_count": len(total),
        "message": ". ".join(msg_parts),
    }
