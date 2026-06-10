"""Парсинг строк ОСВ из CSV, XLSX и Google Sheets."""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

_HEADER_ALIASES: dict[str, str] = {
    "дата": "txn_date",
    "date": "txn_date",
    "маблаги партном": "partner_amount",
    "период оказания услуги": "service_period",
    "выручка": "revenue",
    "выручка - som": "revenue",
    "выручка som": "revenue",
    "revenue": "revenue",
    "расход": "expense",
    "расход - som": "expense",
    "расход som": "expense",
    "expense": "expense",
    "банк": "bank",
    "основание выручка/расход": "basis",
    "основание": "basis",
    "контрагенты": "counterparty",
    "контрагент": "counterparty",
    "телефон": "phone",
    "чрз": "via_person",
    "товар/услуга": "product_service",
    "товар": "product_service",
    "услуга": "product_service",
    "статья": "article",
    "подробно": "detail_category",
    "кратко": "brief_category",
    "ост факт": "balance_hint",
    "som": "balance_hint",
}

_RU_MONTHS = {
    "янв": 1,
    "фев": 2,
    "мар": 3,
    "апр": 4,
    "май": 5,
    "июн": 6,
    "июл": 7,
    "авг": 8,
    "сен": 9,
    "сент": 9,
    "окт": 10,
    "ноя": 11,
    "дек": 12,
}


def normalize_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def parse_decimal(raw: Any) -> Decimal:
    if raw is None:
        return Decimal("0")
    if isinstance(raw, (int, float, Decimal)):
        return Decimal(str(raw))
    s = str(raw).strip().replace("\u00a0", "").replace(" ", "")
    if not s or s in ("-", "—"):
        return Decimal("0")
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def parse_date(raw: Any) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    m2 = re.search(r"(\d{1,2})\s+([а-яa-z]{3,5})\.?", s.lower())
    if m2:
        d = int(m2.group(1))
        mon_key = m2.group(2)[:4]
        for key, mo in _RU_MONTHS.items():
            if mon_key.startswith(key):
                year = date.today().year
                try:
                    return date(year, mo, d)
                except ValueError:
                    return None
    return None


def row_from_mapping(data: dict[str, Any]) -> dict[str, Any] | None:
    txn_date = parse_date(data.get("txn_date"))
    revenue = parse_decimal(data.get("revenue"))
    expense = parse_decimal(data.get("expense"))
    if txn_date is None or (revenue <= 0 and expense <= 0):
        return None
    partner = parse_decimal(data.get("partner_amount"))
    return {
        "txn_date": txn_date,
        "partner_amount": partner if partner > 0 else None,
        "service_period": (str(data.get("service_period") or "").strip() or None),
        "revenue": revenue,
        "expense": expense,
        "bank": (str(data.get("bank") or "").strip() or None),
        "basis": (str(data.get("basis") or "").strip() or None),
        "counterparty": (str(data.get("counterparty") or "").strip() or None),
        "phone": (str(data.get("phone") or "").strip() or None),
        "via_person": (str(data.get("via_person") or "").strip() or None),
        "product_service": (str(data.get("product_service") or "").strip() or None),
        "article": (str(data.get("article") or "").strip() or None),
        "detail_category": (str(data.get("detail_category") or "").strip() or None),
        "brief_category": (str(data.get("brief_category") or "").strip() or None),
    }


def _col_map_from_headers(headers: list[str]) -> dict[int, str]:
    col_map: dict[int, str] = {}
    for i, h in enumerate(headers):
        field = _HEADER_ALIASES.get(normalize_header(h))
        if field and field != "balance_hint":
            col_map[i] = field
    return col_map


def find_osv_header_row(rows: list[list[Any]], *, scan_limit: int = 30) -> tuple[int | None, dict[int, str]]:
    best_idx: int | None = None
    best_map: dict[int, str] = {}
    best_score = 0
    for idx, row in enumerate(rows[:scan_limit]):
        headers = [str(c or "") for c in row]
        col_map = _col_map_from_headers(headers)
        score = len(col_map)
        if "txn_date" in col_map.values():
            score += 2
        if "revenue" in col_map.values() or "expense" in col_map.values():
            score += 2
        if score > best_score:
            best_score = score
            best_idx = idx
            best_map = col_map
    if best_score < 3:
        return None, {}
    return best_idx, best_map


def parse_osv_grid(rows: list[list[Any]]) -> list[dict[str, Any]]:
    header_idx, col_map = find_osv_header_row(rows)
    if header_idx is None or not col_map:
        return []
    out: list[dict[str, Any]] = []
    for line in rows[header_idx + 1 :]:
        if not any(str(c or "").strip() for c in line):
            continue
        data: dict[str, Any] = {}
        for i, field in col_map.items():
            if i < len(line):
                data[field] = line[i]
        parsed = row_from_mapping(data)
        if parsed:
            out.append(parsed)
    return out


def parse_csv_bytes(raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    return parse_osv_grid(rows)


def parse_xlsx_bytes(raw: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    grid = [[c for c in row] for row in rows]
    return parse_osv_grid(grid)
